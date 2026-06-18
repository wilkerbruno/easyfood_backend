# =============================================================
#  EASYFOOD - Rotas do Restaurante
# =============================================================

import bcrypt
from datetime import datetime, date, timedelta
from functools import wraps

from flask import Blueprint, jsonify, request
from flask_jwt_extended import (
    create_access_token, jwt_required,
    get_jwt_identity, get_jwt,
)
from sqlalchemy import func

from backend.models import (
    TableQRCode,
    db, Employee, Restaurant, Order, OrderItem,
    OrderStatusHistory, MenuItem, MenuCategory,
    InventoryItem, InventoryMovement, Review, Payment,
)

restaurant_bp = Blueprint("restaurant", __name__, url_prefix="/api/v1/restaurant")


# ── Helpers ───────────────────────────────────────────────────

def current_employee():
    emp_id = get_jwt_identity()          # vem como string
    return Employee.query.get(int(emp_id))


def require_role(*roles):
    def decorator(f):
        @wraps(f)
        @jwt_required()
        def wrapper(*args, **kwargs):
            claims = get_jwt()
            if claims.get("role") not in roles:
                return jsonify({"error": "Acesso negado"}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator


# ── Auth ──────────────────────────────────────────────────────

@restaurant_bp.post("/auth/login")
def login():
    data     = request.get_json() or {}
    email    = data.get("email", "").strip().lower()
    password = data.get("password", "")

    if not email or not password:
        return jsonify({"error": "E-mail e senha sao obrigatorios"}), 400

    # Busca o funcionário mais recente com este email
    employee = Employee.query.filter_by(email=email)        .order_by(Employee.id.desc()).first()

    if not employee:
        return jsonify({"error": "Credenciais invalidas"}), 401

    if not employee.is_active:
        return jsonify({"error": "Funcionario desativado"}), 401

    try:
        pwd_ok = bcrypt.checkpw(
            password.encode("utf-8"),
            employee.password_hash.encode("utf-8")
        )
    except Exception as e:
        return jsonify({"error": f"Erro ao verificar senha: {str(e)}"}), 500

    if not pwd_ok:
        return jsonify({"error": "Credenciais invalidas"}), 401

    token = create_access_token(
        identity=str(employee.id),
        additional_claims={
            "role":          employee.role,
            "restaurant_id": employee.restaurant_id,
        },
    )
    return jsonify({"access_token": token, "employee": employee.to_dict()})


# ── Dashboard ─────────────────────────────────────────────────

@restaurant_bp.get("/dashboard")
@jwt_required()
def dashboard():
    emp   = current_employee()
    rid   = emp.restaurant_id
    today = datetime.utcnow().date()
    start = datetime.combine(today, datetime.min.time())

    total_sales = (
        db.session.query(func.sum(Order.total))
        .filter(Order.restaurant_id == rid,
                Order.payment_status == "paid",
                Order.created_at >= start)
        .scalar() or 0
    )
    orders_today = Order.query.filter(
        Order.restaurant_id == rid, Order.created_at >= start).count()
    pending   = Order.query.filter_by(restaurant_id=rid, status="pending").count()
    preparing = Order.query.filter_by(restaurant_id=rid, status="preparing").count()
    low_stock = InventoryItem.query.filter(
        InventoryItem.restaurant_id == rid,
        InventoryItem.is_active == True,
        InventoryItem.quantity   <= InventoryItem.min_quantity,
    ).count()
    expiry_warning = InventoryItem.query.filter(
        InventoryItem.restaurant_id == rid,
        InventoryItem.is_active     == True,
        InventoryItem.expiry_date   != None,
        InventoryItem.expiry_date   <= date.today() + timedelta(days=7),
        InventoryItem.expiry_date   >= date.today(),
    ).count()

    return jsonify({
        "total_sales_today": float(total_sales),
        "orders_today":      orders_today,
        "pending_orders":    pending,
        "preparing_orders":  preparing,
        "low_stock_alerts":  low_stock,
        "expiry_alerts":     expiry_warning,
    })


# ── Pedidos ───────────────────────────────────────────────────

@restaurant_bp.get("/orders")
@jwt_required()
def list_orders():
    emp    = current_employee()
    status = request.args.get("status", "").strip()

    query = Order.query.filter_by(restaurant_id=emp.restaurant_id)
    if status:
        query = query.filter_by(status=status)

    orders = query.order_by(Order.created_at.desc()).limit(200).all()
    return jsonify([o.to_dict(include_items=True) for o in orders])


@restaurant_bp.patch("/orders/<int:order_id>/status")
@jwt_required()
def update_order_status(order_id):
    emp   = current_employee()
    order = Order.query.filter_by(
        id=order_id, restaurant_id=emp.restaurant_id
    ).first_or_404()

    data       = request.get_json() or {}
    new_status = data.get("status", "")
    notes      = data.get("notes", "")

    valid = ["confirmed", "preparing", "ready", "delivered", "cancelled"]
    if new_status not in valid:
        return jsonify({"error": f"Status invalido. Use: {valid}"}), 400

    order.status = new_status
    if new_status == "delivered":
        order.delivered_at = datetime.utcnow()

    history = OrderStatusHistory(
        order_id=order.id, status=new_status,
        changed_by=emp.id, notes=notes,
    )
    db.session.add(history)
    db.session.commit()

    # Notifica o cliente sobre a mudanca de status
    try:
        from backend.models import Customer
        from backend.firebase_notify import notify_order_status
        customer = db.session.get(Customer, order.customer_id)
        if customer:
            notify_order_status(customer, order)
    except Exception as e:
        print(f"[NOTIFY] Erro ao notificar status: {e}")

    return jsonify(order.to_dict(include_items=True))


# ── Cardapio ─────────────────────────────────────────────────

@restaurant_bp.get("/menu/categories")
@jwt_required()
def list_categories():
    emp  = current_employee()
    cats = MenuCategory.query.filter_by(
        restaurant_id=emp.restaurant_id, is_active=True
    ).order_by(MenuCategory.display_order).all()
    return jsonify([c.to_dict() for c in cats])


@restaurant_bp.get("/menu")
@jwt_required()
def get_menu():
    emp = current_employee()
    from backend.models import MenuCategory, MenuItem
    categories = []
    cats = MenuCategory.query.filter_by(
        restaurant_id=emp.restaurant_id, is_active=True
    ).order_by(MenuCategory.display_order).all()
    for cat in cats:
        items = MenuItem.query.filter_by(
            category_id=cat.id, is_active=True
        ).all()
        cat_dict = cat.to_dict()
        cat_dict["items"] = [i.to_dict() for i in items]
        categories.append(cat_dict)

    # Itens sem categoria
    no_cat = MenuItem.query.filter_by(
        restaurant_id=emp.restaurant_id, category_id=None, is_active=True
    ).all()
    if no_cat:
        categories.append({
            "id": None, "name": "Geral", "display_order": 999,
            "items": [i.to_dict() for i in no_cat]
        })

    return jsonify({"categories": categories, "total": sum(len(c["items"]) for c in categories)})


def _resolve_category_id(emp, data):
    """
    Resolve o category_id a partir de category_id OU category_name.
    Se category_name nao existir ainda para o restaurante, cria automaticamente.
    """
    if data.get("category_id"):
        return int(data["category_id"])

    category_name = (data.get("category_name") or "").strip()
    if not category_name:
        return None

    cat = MenuCategory.query.filter_by(
        restaurant_id=emp.restaurant_id, name=category_name
    ).first()
    if cat:
        return cat.id

    # Cria a categoria nova automaticamente
    max_order = db.session.query(db.func.max(MenuCategory.display_order)).filter_by(
        restaurant_id=emp.restaurant_id
    ).scalar() or 0
    cat = MenuCategory(
        restaurant_id=emp.restaurant_id,
        name=category_name,
        display_order=max_order + 1,
    )
    db.session.add(cat)
    db.session.flush()  # garante que cat.id existe antes de usar
    return cat.id


@restaurant_bp.post("/menu/items")
@require_role("admin")
def create_menu_item():
    emp  = current_employee()
    data = request.get_json() or {}
    category_id = _resolve_category_id(emp, data)
    item = MenuItem(
        restaurant_id    = emp.restaurant_id,
        category_id      = category_id,
        name             = data["name"],
        description      = data.get("description"),
        price            = data["price"],
        image_url        = data.get("image_url"),
        preparation_time = data.get("preparation_time"),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify(item.to_dict()), 201


@restaurant_bp.put("/menu/items/<int:item_id>")
@require_role("admin")
def update_menu_item(item_id):
    emp  = current_employee()
    item = MenuItem.query.filter_by(
        id=item_id, restaurant_id=emp.restaurant_id
    ).first_or_404()
    data = request.get_json() or {}
    if data.get("category_name") and not data.get("category_id"):
        data["category_id"] = _resolve_category_id(emp, data)
    for field in ["name","description","price","image_url","preparation_time","is_available","category_id"]:
        if field in data:
            setattr(item, field, data[field])
    db.session.commit()
    return jsonify(item.to_dict())


@restaurant_bp.delete("/menu/items/<int:item_id>")
@require_role("admin")
def delete_menu_item(item_id):
    emp  = current_employee()
    item = MenuItem.query.filter_by(
        id=item_id, restaurant_id=emp.restaurant_id
    ).first_or_404()
    item.is_active = False
    db.session.commit()
    return jsonify({"message": "Item removido"})


# ── Estoque ───────────────────────────────────────────────────

@restaurant_bp.get("/inventory")
@jwt_required()
def list_inventory():
    emp   = current_employee()
    items = InventoryItem.query.filter_by(
        restaurant_id=emp.restaurant_id, is_active=True
    ).order_by(InventoryItem.name).all()
    return jsonify([i.to_dict() for i in items])


@restaurant_bp.post("/inventory")
@require_role("admin")
def create_inventory_item():
    emp  = current_employee()
    data = request.get_json() or {}

    expiry = None
    if data.get("expiry_date"):
        try:
            expiry = datetime.strptime(data["expiry_date"], "%Y-%m-%d").date()
        except ValueError:
            pass

    item = InventoryItem(
        restaurant_id = emp.restaurant_id,
        name          = data["name"],
        category      = data.get("category"),
        unit_type     = data.get("unit_type", "unit"),
        quantity      = float(data.get("quantity", 0)),
        min_quantity  = float(data.get("min_quantity", 0)),
        cost_price    = data.get("cost_price"),
        supplier      = data.get("supplier"),
        expiry_date   = expiry,
        batch_number  = data.get("batch_number"),
        location      = data.get("location"),
        notes         = data.get("notes"),
    )
    db.session.add(item)
    db.session.flush()

    if float(data.get("quantity", 0)) > 0:
        mov = InventoryMovement(
            inventory_item_id = item.id,
            restaurant_id     = emp.restaurant_id,
            employee_id       = emp.id,
            movement_type     = "in",
            quantity          = item.quantity,
            quantity_before   = 0,
            quantity_after    = item.quantity,
            reason            = "Cadastro inicial",
        )
        db.session.add(mov)

    db.session.commit()
    return jsonify(item.to_dict()), 201


@restaurant_bp.put("/inventory/<int:item_id>")
@require_role("admin")
def update_inventory_item(item_id):
    emp  = current_employee()
    item = InventoryItem.query.filter_by(
        id=item_id, restaurant_id=emp.restaurant_id, is_active=True
    ).first_or_404()
    data = request.get_json() or {}
    for field in ["name","category","unit_type","min_quantity","cost_price","supplier","batch_number","location","notes"]:
        if field in data:
            setattr(item, field, data[field])
    if data.get("expiry_date"):
        try:
            item.expiry_date = datetime.strptime(data["expiry_date"], "%Y-%m-%d").date()
        except ValueError:
            pass
    db.session.commit()
    return jsonify(item.to_dict())


@restaurant_bp.post("/inventory/<int:item_id>/movement")
@jwt_required()
def add_movement(item_id):
    emp  = current_employee()
    item = InventoryItem.query.filter_by(
        id=item_id, restaurant_id=emp.restaurant_id, is_active=True
    ).first_or_404()

    data     = request.get_json() or {}
    mov_type = data.get("movement_type")
    qty      = float(data.get("quantity", 0))
    reason   = data.get("reason", "")

    if mov_type not in ("in", "out", "adjustment", "waste"):
        return jsonify({"error": "movement_type invalido"}), 400
    if qty <= 0:
        return jsonify({"error": "Quantidade deve ser maior que zero"}), 400

    qty_before = float(item.quantity)
    if mov_type == "in":
        qty_after = qty_before + qty
    elif mov_type in ("out", "waste"):
        qty_after = max(0, qty_before - qty)
    else:
        qty_after = qty   # adjustment = valor absoluto

    mov = InventoryMovement(
        inventory_item_id = item.id,
        restaurant_id     = emp.restaurant_id,
        employee_id       = emp.id,
        movement_type     = mov_type,
        quantity          = qty,
        quantity_before   = qty_before,
        quantity_after    = qty_after,
        reason            = reason,
    )
    item.quantity = qty_after
    db.session.add(mov)
    db.session.commit()
    return jsonify({"message": "Movimentacao registrada", "item": item.to_dict()})


@restaurant_bp.get("/inventory/<int:item_id>/history")
@jwt_required()
def inventory_history(item_id):
    emp  = current_employee()
    item = InventoryItem.query.filter_by(
        id=item_id, restaurant_id=emp.restaurant_id
    ).first_or_404()
    movements = item.movements.order_by(InventoryMovement.created_at.desc()).limit(50).all()
    return jsonify([{
        "id": m.id, "movement_type": m.movement_type,
        "quantity": float(m.quantity),
        "quantity_before": float(m.quantity_before),
        "quantity_after":  float(m.quantity_after),
        "reason":     m.reason,
        "created_at": m.created_at.isoformat(),
    } for m in movements])


# ── Relatorios ────────────────────────────────────────────────

@restaurant_bp.get("/reports/sales")
@require_role("admin")
def sales_report():
    emp       = current_employee()
    date_from = request.args.get("date_from")
    date_to   = request.args.get("date_to")

    query = Order.query.filter_by(
        restaurant_id=emp.restaurant_id, payment_status="paid"
    )
    if date_from:
        try:
            query = query.filter(Order.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Order.created_at <= datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass

    orders = query.all()
    total  = sum(float(o.total) for o in orders)

    item_sales = (
        db.session.query(
            MenuItem.name,
            func.sum(OrderItem.quantity).label("qty_sold"),
            func.sum(OrderItem.subtotal).label("revenue"),
        )
        .join(OrderItem, OrderItem.menu_item_id == MenuItem.id)
        .join(Order, Order.id == OrderItem.order_id)
        .filter(
            Order.restaurant_id  == emp.restaurant_id,
            Order.payment_status == "paid",
        )
        .group_by(MenuItem.id, MenuItem.name)
        .order_by(func.sum(OrderItem.quantity).desc())
        .limit(10)
        .all()
    )

    return jsonify({
        "period":        {"from": date_from, "to": date_to},
        "total_orders":  len(orders),
        "total_revenue": total,
        "top_items": [
            {"name": r.name, "qty_sold": int(r.qty_sold), "revenue": float(r.revenue)}
            for r in item_sales
        ],
    })


# ── QR Codes de Mesa ──────────────────────────────────────────

@restaurant_bp.get("/qrcodes")
@jwt_required()
def list_qrcodes():
    emp  = current_employee()
    rest = Restaurant.query.get(emp.restaurant_id)
    # QR codes sao compartilhados por toda a praca de alimentacao
    qrs  = TableQRCode.query.filter_by(
        food_court_id=rest.food_court_id, is_active=True
    ).order_by(TableQRCode.table_number).all()
    return jsonify([q.to_dict() for q in qrs])


@restaurant_bp.post("/qrcodes")
@require_role("admin")
def create_qrcode():
    import secrets
    emp  = current_employee()
    rest = Restaurant.query.get(emp.restaurant_id)
    data = request.get_json() or {}

    table  = data.get("table_number", "").strip()
    label  = data.get("label", f"Mesa {table}").strip()
    qty    = min(int(data.get("quantity", 1)), 50)

    if not table:
        return jsonify({"error": "Numero da mesa e obrigatorio"}), 400

    created = []
    for i in range(qty):
        suffix = f"-{i+1}" if qty > 1 else ""
        # Verifica se ja existe uma mesa com esse numero na praca
        existing = TableQRCode.query.filter_by(
            food_court_id=rest.food_court_id,
            table_number=f"{table}{suffix}",
            is_active=True
        ).first()
        if existing:
            created.append(existing)
            continue
        qr = TableQRCode(
            food_court_id = rest.food_court_id,
            restaurant_id = None,  # compartilhado por toda a praca
            table_number  = f"{table}{suffix}",
            qr_token      = secrets.token_hex(16),
            label         = f"{label}{suffix}",
        )
        db.session.add(qr)
        created.append(qr)

    db.session.commit()
    return jsonify([q.to_dict() for q in created]), 201


@restaurant_bp.delete("/qrcodes/<int:qr_id>")
@require_role("admin")
def delete_qrcode(qr_id):
    emp  = current_employee()
    rest = Restaurant.query.get(emp.restaurant_id)
    # Permite remover apenas QR codes da mesma praca
    qr = TableQRCode.query.filter_by(
        id=qr_id, food_court_id=rest.food_court_id
    ).first_or_404()
    qr.is_active = False
    db.session.commit()
    return jsonify({"message": "QR Code removido"})


# ══════════════════════════════════════════════════════════════
# UPLOAD EXCEL - CARDÁPIO
# ══════════════════════════════════════════════════════════════

@restaurant_bp.post("/menu/import")
@jwt_required()
def import_menu_excel(emp):
    """Importa cardápio via planilha Excel."""
    import io
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl não instalado"}), 500

    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Arquivo deve ser .xlsx ou .xls"}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
    except Exception as e:
        return jsonify({"error": f"Erro ao ler planilha: {e}"}), 400

    created = 0
    updated = 0
    errors  = []

    for sheet in wb.worksheets:
        category_name = sheet.title.strip() or "Geral"

        # Cria ou busca a categoria
        cat = MenuCategory.query.filter_by(
            restaurant_id=emp.restaurant_id,
            name=category_name
        ).first()
        if not cat:
            cat = MenuCategory(restaurant_id=emp.restaurant_id, name=category_name)
            db.session.add(cat)
            db.session.flush()

        # Lê as linhas (ignora cabeçalho)
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            try:
                def get(col):
                    try:
                        idx = headers.index(col)
                        return row[idx] if idx < len(row) else None
                    except ValueError:
                        return None

                name  = str(get("nome") or get("name") or row[0] or "").strip()
                price = get("preco") or get("preço") or get("price") or get("valor") or 0
                desc  = get("descricao") or get("descrição") or get("description") or ""
                prep  = get("tempo") or get("preparo") or get("preparation_time") or 0
                avail = get("disponivel") or get("disponível") or get("available") or True

                if not name:
                    continue

                try:
                    price = float(str(price).replace("R$","").replace(",",".").strip())
                except:
                    price = 0.0

                # Verifica se já existe
                existing = MenuItem.query.filter_by(
                    restaurant_id=emp.restaurant_id,
                    name=name,
                    is_active=True
                ).first()

                if existing:
                    existing.price       = price
                    existing.description = str(desc) if desc else existing.description
                    existing.category_id = cat.id
                    updated += 1
                else:
                    item = MenuItem(
                        restaurant_id    = emp.restaurant_id,
                        category_id      = cat.id,
                        name             = name,
                        description      = str(desc) if desc else None,
                        price            = price,
                        preparation_time = int(prep) if prep else None,
                        is_available     = bool(avail) if avail != "" else True,
                    )
                    db.session.add(item)
                    created += 1

            except Exception as e:
                errors.append(f"Linha {row_idx}: {e}")

    db.session.commit()
    return jsonify({
        "message": f"Importação concluída! {created} criados, {updated} atualizados.",
        "created": created, "updated": updated, "errors": errors
    }), 200


@restaurant_bp.get("/menu/template")
@jwt_required()
def download_menu_template(emp):
    """Baixa template Excel para importação de cardápio."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl não instalado"}), 500

    wb = openpyxl.Workbook()

    # Aba 1 — Lanches
    for sheet_name in ["Lanches", "Bebidas", "Sobremesas"]:
        ws = wb.create_sheet(sheet_name)
        headers = ["nome", "preco", "descricao", "tempo", "disponivel"]
        header_fill = PatternFill("solid", fgColor="FF6B35")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h.upper())
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 20

        # Exemplos
        examples = [
            ["X-Burguer",  15.90, "Hamburguer artesanal com queijo",  15, True],
            ["X-Salada",   17.90, "Hamburguer com alface e tomate",   15, True],
            ["Batata Frita", 8.90, "Porção de batata frita crocante", 10, True],
        ]
        if sheet_name == "Bebidas":
            examples = [
                ["Coca-Cola 350ml", 5.00, "Refrigerante gelado", 2, True],
                ["Suco de Laranja", 7.00, "Suco natural 300ml",  3, True],
            ]
        elif sheet_name == "Sobremesas":
            examples = [
                ["Sorvete",    6.00, "Bola de sorvete de creme", 3, True],
                ["Pudim",      7.00, "Pudim de leite condensado", 2, True],
            ]
        for r, ex in enumerate(examples, 2):
            for c, val in enumerate(ex, 1):
                ws.cell(row=r, column=c, value=val)

    # Remove aba padrão
    del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="template_cardapio.xlsx"
    )


# ══════════════════════════════════════════════════════════════
# UPLOAD EXCEL - ESTOQUE
# ══════════════════════════════════════════════════════════════

@restaurant_bp.post("/inventory/import")
@jwt_required()
def import_inventory_excel(emp):
    """Importa estoque via planilha Excel."""
    import io
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl não instalado"}), 500

    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Arquivo deve ser .xlsx ou .xls"}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"Erro ao ler planilha: {e}"}), 400

    created = 0
    updated = 0
    errors  = []

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            def get(col):
                try:
                    idx = headers.index(col)
                    return row[idx] if idx < len(row) else None
                except ValueError:
                    return None

            name      = str(get("nome") or get("name") or row[0] or "").strip()
            qty       = get("quantidade") or get("quantity") or get("qtd") or 0
            min_qty   = get("minimo") or get("mínimo") or get("min") or 0
            unit      = get("unidade") or get("unit") or "unit"
            cost      = get("custo") or get("cost") or get("preco") or 0
            category  = get("categoria") or get("category") or ""
            supplier  = get("fornecedor") or get("supplier") or ""

            if not name:
                continue

            valid_units = ["unit","kg","g","l","ml"]
            unit = str(unit).lower().strip() if unit else "unit"
            if unit not in valid_units:
                unit = "unit"

            try: qty  = float(str(qty).replace(",","."))
            except: qty = 0
            try: min_qty = float(str(min_qty).replace(",","."))
            except: min_qty = 0
            try: cost = float(str(cost).replace("R$","").replace(",",".").strip())
            except: cost = 0

            existing = InventoryItem.query.filter_by(
                restaurant_id=emp.restaurant_id,
                name=name,
                is_active=True
            ).first()

            if existing:
                existing.quantity     = qty
                existing.min_quantity = min_qty
                existing.unit_type    = unit
                existing.cost_price   = cost
                if category: existing.category = str(category)
                if supplier: existing.supplier  = str(supplier)
                updated += 1
            else:
                item = InventoryItem(
                    restaurant_id = emp.restaurant_id,
                    name          = name,
                    quantity      = qty,
                    min_quantity  = min_qty,
                    unit_type     = unit,
                    cost_price    = cost,
                    category      = str(category) if category else None,
                    supplier      = str(supplier) if supplier else None,
                )
                db.session.add(item)
                created += 1

        except Exception as e:
            errors.append(f"Linha {row_idx}: {e}")

    db.session.commit()
    return jsonify({
        "message": f"Importação concluída! {created} criados, {updated} atualizados.",
        "created": created, "updated": updated, "errors": errors
    }), 200


@restaurant_bp.get("/inventory/template")
@jwt_required()
def download_inventory_template(emp):
    """Baixa template Excel para importação de estoque."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl não instalado"}), 500

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Estoque"

    headers = ["nome","quantidade","minimo","unidade","custo","categoria","fornecedor"]
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h.upper())
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    examples = [
        ["Pão de Hamburguer",  50,  20, "unit", 0.80, "Panificados",   "Padaria Central"],
        ["Carne Bovina",        5,   2, "kg",   35.00, "Proteínas",    "Frigorífico ABC"],
        ["Queijo Cheddar",      3,   1, "kg",   28.00, "Laticínios",   "Laticínio XYZ"],
        ["Coca-Cola 350ml",   100,  30, "unit",  2.50, "Bebidas",      "Distribuidora EFG"],
        ["Óleo de Soja",        2, 0.5, "l",     6.00, "Ingredientes", "Mercado Local"],
    ]
    for r, ex in enumerate(examples, 2):
        for c, val in enumerate(ex, 1):
            ws.cell(row=r, column=c, value=val)

    # Instruções
    ws2 = wb.create_sheet("Instruções")
    instrucoes = [
        ["CAMPO",       "DESCRIÇÃO",                    "EXEMPLO"],
        ["nome",        "Nome do item (obrigatório)",    "Pão de Hamburguer"],
        ["quantidade",  "Quantidade atual em estoque",   "50"],
        ["minimo",      "Quantidade mínima (alerta)",    "20"],
        ["unidade",     "unit / kg / g / l / ml",        "unit"],
        ["custo",       "Preço de custo unitário",        "0.80"],
        ["categoria",   "Categoria do item",             "Panificados"],
        ["fornecedor",  "Nome do fornecedor",            "Padaria Central"],
    ]
    for r, row in enumerate(instrucoes, 1):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)
            ws2.column_dimensions[cell.column_letter].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="template_estoque.xlsx"
    )
